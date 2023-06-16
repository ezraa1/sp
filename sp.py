import itertools
import openpyxl

matches = ['1', 'x', '2']
all_combinations = list(itertools.product(matches, repeat=13))
filtered_combinations = []
for comb in all_combinations:
    if 'x' not in comb:
        continue
    if '1' not in comb:
        continue
    if '2' not in comb:
        continue
    if all(result == '1' for result in comb[:10]):
        continue
    if all(result == '2' for result in comb[:10]):
        continue
    if all(result == 'x' for result in comb[:10]):
        continue
    if all(result == 'x' for result in comb[-5:]):
        continue
    if all(result == '1' for result in comb[-5:]):
        continue
    if all(result == '2' for result in comb[-5:]):
        continue
    if any(comb[i] == comb[i+1] == comb[i+2] == comb[i+3] == comb[i+4] != 'x' for i in range(8)):
        continue
    if '1' not in comb[6:]:
        continue
    if '2' not in comb[6:] or 'x' not in comb[6:]:
        continue
    filtered_combinations.append(comb)

# Split combinations into chunks of 1048575 (maximum number of rows per worksheet - 1)
chunk_size = 1048575
combinations_chunks = [
    filtered_combinations[i:i+chunk_size]
    for i in range(0, len(filtered_combinations), chunk_size)
]

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Save combinations in separate worksheets
for i, combinations_chunk in enumerate(combinations_chunks):
    sheet = workbook.create_sheet(title=f'Sheet{i+1}')

    for j, combination in enumerate(combinations_chunk):
        sheet.cell(row=j+1, column=1, value=''.join(combination))

# Remove the default sheet created with the workbook
workbook.remove(workbook['Sheet'])

# Save the workbook
workbook.save('winning_combinations.xlsx')
print("Combinations saved to winning_combinations.xlsx")